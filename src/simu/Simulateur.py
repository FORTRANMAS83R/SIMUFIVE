import sys
import os
import argparse
import matplotlib.pyplot as plt
import warnings
import time


# Ajouter le répertoire parent à sys.path
parent_directory = os.path.abspath(os.path.join(os.path.dirname(__file__), '..','..'))
if parent_directory not in sys.path:
    sys.path.append(parent_directory)
import src.config.config as config
from src.simu.Sport import Sport
from src.simu.Bar import Bar
from src.simu.Semaine import Semaine
from openpyxl import Workbook

class Simulateur: 
    def __init__(self,config, nb_realisations): 
        self.nb_realisations = nb_realisations
        self.realisations = []
        self.config = config

    def run(self, parallel = False):
        if parallel:
            # Parallel computing
            pass
        else:
            for i in range(self.nb_realisations):
                current_realisation = Realisation(self.config)
                current_realisation.run()
                self.realisations.append(current_realisation)

    def mean(self):
        mean = Realisation(self.config)
        for i in range(mean.duree): 
            s = Semaine()
            s.mean([realisation.semaines[i] for realisation in self.realisations])
            mean.semaines.append(s)
        return mean

    def toXlsx(self, fileName): 
        # Put simulation into an excel file
        # First sheet is for the global simulation (raw data)
        # Second sheet is for the mean CA per week, month, year
        wb = Workbook()

        realisation = self.mean()

        # First sheet
        ws = wb.active
        ws.title = "Simulation"
        ws.append(["Semaine", "Revenus Five", "Revenus Beach", "Revenus Padel", "Revenus Bar", "Reservations Five HC", "Reservations Five HP", "Reservations Beach HC", "Reservations Beach HP", "Reservations Padel HC", "Reservations Padel HP", "Affluence Bar", "% Five", "% Beach", "% Padel", "% Bar"])
        for i in range(len(realisation.semaines)):
            semaine = realisation.semaines[i]
            ws.append([i, semaine.five.revenu, semaine.beach.revenu, semaine.padel.revenu, semaine.bar.revenus, semaine.five.freq.res_hc, semaine.five.freq.res_hp, semaine.beach.freq.res_hc, semaine.beach.freq.res_hp, semaine.padel.freq.res_hc, semaine.padel.freq.res_hp, semaine.bar.affluence, semaine.repartition["Five"], semaine.repartition["Beach"], semaine.repartition["Padel"], semaine.repartition["Bar"]])
        
        # Second sheet
        ws = wb.create_sheet(title="Moyennes")
        ws.append(["Période", "Revenus Five", "Revenus Beach", "Revenus Padel", "Revenus Bar"])
        ws.append(["Semaine", sum([semaine.five.revenu for semaine in realisation.semaines])/len(realisation.semaines), sum([semaine.beach.revenu for semaine in realisation.semaines])/len(realisation.semaines), sum([semaine.padel.revenu for semaine in realisation.semaines])/len(realisation.semaines), sum([semaine.bar.revenus for semaine in realisation.semaines])/len(realisation.semaines)])
        ws.append(["Mois", sum([semaine.five.revenu for semaine in realisation.semaines])/len(realisation.semaines)*4, sum([semaine.beach.revenu for semaine in realisation.semaines])/len(realisation.semaines)*4, sum([semaine.padel.revenu for semaine in realisation.semaines])/len(realisation.semaines)*4, sum([semaine.bar.revenus for semaine in realisation.semaines])/len(realisation.semaines)*4])
        ws.append(["Année", sum([semaine.five.revenu for semaine in realisation.semaines])/len(realisation.semaines)*52, sum([semaine.beach.revenu for semaine in realisation.semaines])/len(realisation.semaines)*52, sum([semaine.padel.revenu for semaine in realisation.semaines])/len(realisation.semaines)*52, sum([semaine.bar.revenus for semaine in realisation.semaines])/len(realisation.semaines)*52])

        # Third sheet
        ws = wb.create_sheet(title="Configurations")
        ws.append(["Sport", "Nombre terrain", "Prix HC", "Prix HP", "Freq Init HC", "Freq Max HC", "Freq Init HP", "Freq Max HP"])
        ws.append(["Five", realisation.Five.nb_terrains, realisation.Five.prix_hc, realisation.Five.prix_hp, realisation.Five.freq.freq_init_hc, realisation.Five.freq.freq_max_hc, realisation.Five.freq.freq_init_hp, realisation.Five.freq.freq_max_hp])
        ws.append(["Beach", realisation.Beach.nb_terrains, realisation.Beach.prix_hc, realisation.Beach.prix_hp, realisation.Beach.freq.freq_init_hc, realisation.Beach.freq.freq_max_hc, realisation.Beach.freq.freq_init_hp, realisation.Beach.freq.freq_max_hp])
        ws.append(["Padel", realisation.Padel.nb_terrains, realisation.Padel.prix_hc, realisation.Padel.prix_hp, realisation.Padel.freq.freq_init_hc, realisation.Padel.freq.freq_max_hc, realisation.Padel.freq.freq_init_hp, realisation.Padel.freq.freq_max_hp])

        # Write file
        wb.save(fileName + ".xlsx")

    def __str__(self):
        c = ''
        for i in range(self.nb_realisations):
            c += f"Realisation {i+1} : \n"
            c += str(self.realisations[i])
        return c


"""
Classe Simulateur, stocke des objets sports et leurs difffférentes copies au cours des semaines. 
Permetant ainsi de simuler l'évolution des sports au cours du temps.
"""
class Realisation: 
    def __init__(self, config): 
        self.duree = config.duree_simu
        self.Five = Sport()
        self.Five.set_config(config.Five)
        self.Beach = Sport()
        self.Beach.set_config(config.Beach)
        self.Padel = Sport()
        self.Padel.set_config(config.Padel)
        self.Bar = Bar()
        self.Bar.set_config(config.Bar)
        self.semaines = []
        self.revenus = 0
        
    

    #Ajoute une semaine à la simulation

    def add(self, semaine):
        self.semaines.append(semaine)
        self.revenus += semaine.five.revenu + semaine.beach.revenu + semaine.padel.revenu + semaine.bar.revenus
    
    #Fait évoluer les sports de la simulation
    
    def evolve(self, i): 
        self.Five.evolve(i)
        self.Beach.evolve(i)
        self.Padel.evolve(i)
        self.Bar.evolve(i)
    
    
    def plot(self):
        n_semaine=[]
        res_five=[]
        res_beach=[]
        res_padel=[]
        CA=[]
        i=0
        for semaine in self.semaines:
            n_semaine.append(i)
            res_five.append(semaine.five.revenu)
            res_beach.append(semaine.beach.revenu)
            res_padel.append(semaine.padel.revenu)
            
            i+=1

        plt.plot(n_semaine,res_five,label="Five")
        plt.plot(n_semaine,res_beach,label="Beach")
        plt.plot(n_semaine,res_padel,label="Padel")
        plt.xlabel("Semaine")
        plt.ylabel("Revenu [€]")
        plt.legend()
        plt.show()

    #Lance la simulation
    def run(self): 
        for i in range(self.duree): 
            curr_semaine = Semaine()  
            self.evolve(i)  
            curr_semaine.add(self.Five.clone(), self.Beach.clone(), self.Padel.clone(), self.Bar.clone())
            self.add(curr_semaine)

    
    def __str__(self):
        c = ''
        for i in range(self.duree):
            c += f"Semaine {i+1} : \n"
            c += str(self.semaines[i]) + "\n"
        return c
        
def start_multi_var(params_var: list, init_val: list, nb_simu: int, step: list, config_path: str, nb_realisations: int) -> list:
    if not isinstance(params_var, list) or not isinstance(init_val, list) or not isinstance(step, list):
        raise ValueError("params_var, init_val and step doivent êter des listes ! start_multi_var(params_var: list, init_val: list, nb_simu: int, step: list, config_path: str) -> list")
    if len(params_var) != len(init_val) or len(params_var) != len(step):
        raise ValueError("params_var, init_val et step doivent avoir la même taille !")
    param_init = [config.stocker_param_init(config_path, param_var) for param_var in params_var]
    simulations = []
    for i in range(nb_simu):
        [config.modifier_parametre(config_path, params_var[j], init_val[j] + i*step[j]) for j in range(len(params_var))]
        simulations.append(start(config_path, nb_realisations))
    return simulations

def start(config_path: str, nb_realisations: int) -> Simulateur:
    cfg = config.init(config_path)
    simu = Simulateur(cfg, nb_realisations)
    simu.run()
    return simu



if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Script pour traiter un fichier.')
    parser.add_argument('file_name', type=str, help='Le nom du fichier à traiter')
    parser.add_argument('--xlsx', type=str, help='Le nom du fichier excel de sortie (sans extension)')
    parser.add_argument('--param_var', nargs = '+', help='Lancer une simulation avec un ou plusieurs triplets de paramètres variables (nom_param, val_init, step)')
    parser.add_argument("--nb_simu", type=int, help="Le nombre de simulations à lancer (inutile si --param-var n'est utilisé)")
    parser.add_argument("-time", action="store_true", help="Afficher le temps d'exécution")
    parser.add_argument("--nb_realisations", type=int, help="Le nombre de réalisations à effectuer")
    args = parser.parse_args()
    print("Starting simulation...")
    if(args.time):
        start_time = time.time()
    if(args.nb_realisations):
        nb_realisations = args.nb_realisations
    else:
        nb_realisations = 1
        warnings.warn("Le nombre de réalisations n'a pas été spécifié, par défaut nb_realisations = 1")
    if(args.param_var):
        if(len(args.param_var)%3 != 0):
            raise ValueError("Le nombre d'arguments pour --param_var doit être un multiple de 3 !(nom_param, val_init, step)")
        params_var = []
        val_init = []
        step = []
        for i in range(0, len(args.param_var), 3):
            params_var.append(args.param_var[i])    
            val_init.append(float(args.param_var[i+1]))
            step.append(float(args.param_var[i+2]))
        if(args.nb_simu):
            nb_simu = args.nb_simu
        else:
            nb_simu = 1
            warnings.warn("Le nombre de simulations n'a pas été spécifié, par défaut nb_simu = 1")
        
        simulations = start_multi_var(params_var, val_init, nb_simu, step, args.file_name, nb_realisations)
    else: 
        if(args.nb_simu):
            warnings.warn("Attention : --nb_simu est ignoré car --param-var n’est pas actif")
        simu = start(args.file_name, nb_realisations)
        #simu.plot()
        print(str(simu.mean()))
        if(args.xlsx):
            simu.toXlsx(args.xlsx)
    if(args.time):
        print("Execution time : %s seconds" % (round((time.time() - start_time),1)))
    print("Simulation done !")